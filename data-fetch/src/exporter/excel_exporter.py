"""
Excel exporter for the data-fetch framework.
Exports DataFrames to Excel with optional metadata.
"""

from pathlib import Path
from typing import Optional, Dict, Any, Tuple
from datetime import datetime
import io
import tempfile

import pandas as pd

from ..utils.logger import get_logger
from ..utils.io_utils import get_output_path, ensure_dir, timestamp_now


class ExcelExporter:
    """
    Exporter for saving DataFrames to Excel files.
    Supports single-sheet export with optional metadata sheet.
    """
    
    def __init__(
        self,
        output_dir: Optional[Path] = None,
        include_metadata: bool = True,
        date_format: str = "YYYY-MM-DD",
    ):
        """
        Initialize the exporter.
        
        Args:
            output_dir: Directory for output files
            include_metadata: Whether to include a metadata sheet
            date_format: Excel date format
        """
        self.output_dir = output_dir
        self.include_metadata = include_metadata
        self.date_format = date_format
        self.logger = get_logger()
    
    def export(
        self,
        df: pd.DataFrame,
        filename: Optional[str] = None,
        site_id: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None,
        sheet_name: str = "Data",
    ) -> Path:
        """
        Export a DataFrame to Excel.
        
        Args:
            df: DataFrame to export
            filename: Output filename (auto-generated if not provided)
            site_id: Site identifier for organizing files
            metadata: Additional metadata to include
            sheet_name: Name of the main data sheet
        
        Returns:
            Path to the created Excel file
        """
        # Generate filename if not provided
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            site_part = f"{site_id}_" if site_id else ""
            filename = f"{site_part}data_{timestamp}.xlsx"
        
        # Ensure .xlsx extension
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
        
        # Get output path
        if self.output_dir:
            output_path = self.output_dir / filename
            ensure_dir(self.output_dir)
        else:
            output_path = get_output_path(filename, "excel", site_id)
        
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Write main data sheet
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Format the worksheet
            worksheet = writer.sheets[sheet_name]
            self._format_worksheet(worksheet, df)
            
            # Write metadata sheet if requested
            if self.include_metadata:
                meta_df = self._create_metadata_df(df, metadata, site_id)
                meta_df.to_excel(writer, sheet_name="Metadata", index=False)
        
        self.logger.info(f"Exported {len(df)} rows to {output_path}")
        return output_path
    
    def export_to_bytes(
        self,
        df: pd.DataFrame,
        filename: Optional[str] = None,
        site_id: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None,
        sheet_name: str = "Data",
    ) -> Tuple[bytes, str]:
        """
        Export a DataFrame to Excel in memory (for cloud/streamlit use).
        
        Args:
            df: DataFrame to export
            filename: Output filename (auto-generated if not provided)
            site_id: Site identifier for organizing files
            metadata: Additional metadata to include
            sheet_name: Name of the main data sheet
        
        Returns:
            Tuple of (excel_bytes, filename)
        """
        # Generate filename if not provided
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            site_part = f"{site_id}_" if site_id else ""
            filename = f"{site_part}data_{timestamp}.xlsx"
        
        # Ensure .xlsx extension
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
        
        # Use temporary file or BytesIO for cloud environments
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            tmp_path = Path(tmp_file.name)
        
        try:
            # Create Excel writer
            with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
                # Write main data sheet
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Format the worksheet
                worksheet = writer.sheets[sheet_name]
                self._format_worksheet(worksheet, df)
                
                # Write metadata sheet if requested
                if self.include_metadata:
                    meta_df = self._create_metadata_df(df, metadata, site_id)
                    meta_df.to_excel(writer, sheet_name="Metadata", index=False)
            
            # Read file as bytes
            with open(tmp_path, "rb") as f:
                excel_bytes = f.read()
            
            # Clean up temp file
            tmp_path.unlink()
            
            self.logger.info(f"Exported {len(df)} rows to memory ({len(excel_bytes)} bytes)")
            return excel_bytes, filename
            
        except Exception as e:
            # Clean up on error
            if tmp_path.exists():
                tmp_path.unlink()
            raise e
    
    def export_multiple(
        self,
        dataframes: Dict[str, pd.DataFrame],
        filename: str,
        site_id: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None,
    ) -> Path:
        """
        Export multiple DataFrames to different sheets.
        
        Args:
            dataframes: Dictionary mapping sheet names to DataFrames
            filename: Output filename
            site_id: Site identifier
            metadata: Additional metadata
        
        Returns:
            Path to the created Excel file
        """
        # Ensure .xlsx extension
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
        
        # Get output path
        if self.output_dir:
            output_path = self.output_dir / filename
            ensure_dir(self.output_dir)
        else:
            output_path = get_output_path(filename, "excel", site_id)
        
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            total_rows = 0
            
            for sheet_name, df in dataframes.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]
                self._format_worksheet(worksheet, df)
                total_rows += len(df)
            
            # Write metadata sheet
            if self.include_metadata:
                first_df = list(dataframes.values())[0] if dataframes else pd.DataFrame()
                meta_df = self._create_metadata_df(first_df, metadata, site_id)
                meta_df["sheets"] = ", ".join(dataframes.keys())
                meta_df["total_rows"] = total_rows
                meta_df.to_excel(writer, sheet_name="Metadata", index=False)
        
        self.logger.info(f"Exported {len(dataframes)} sheets to {output_path}")
        return output_path
    
    def _format_worksheet(self, worksheet, df: pd.DataFrame):
        """Apply formatting to a worksheet."""
        from openpyxl.utils import get_column_letter
        
        # Auto-fit column widths
        for idx, col in enumerate(df.columns, 1):
            max_length = max(
                len(str(col)),
                df[col].astype(str).str.len().max() if len(df) > 0 else 0
            )
            # Cap width at 50 characters
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
        
        # Format header row
        from openpyxl.styles import Font, PatternFill
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Freeze header row
        worksheet.freeze_panes = "A2"
    
    def _create_metadata_df(
        self,
        df: pd.DataFrame,
        metadata: Optional[Dict[str, Any]],
        site_id: Optional[str],
    ) -> pd.DataFrame:
        """Create a metadata DataFrame."""
        meta_data = {
            "property": [],
            "value": [],
        }
        
        # Basic info
        meta_data["property"].append("Generated At")
        meta_data["value"].append(timestamp_now())
        
        meta_data["property"].append("Row Count")
        meta_data["value"].append(len(df))
        
        meta_data["property"].append("Column Count")
        meta_data["value"].append(len(df.columns))
        
        meta_data["property"].append("Columns")
        meta_data["value"].append(", ".join(df.columns))
        
        if site_id:
            meta_data["property"].append("Site ID")
            meta_data["value"].append(site_id)
        
        # Date range if available
        date_cols = [c for c in df.columns if "date" in c.lower()]
        if date_cols and len(df) > 0:
            date_col = date_cols[0]
            try:
                dates = pd.to_datetime(df[date_col])
                meta_data["property"].append("Date Range Start")
                meta_data["value"].append(str(dates.min()))
                meta_data["property"].append("Date Range End")
                meta_data["value"].append(str(dates.max()))
            except Exception:
                pass
        
        # Add custom metadata
        if metadata:
            for key, value in metadata.items():
                meta_data["property"].append(key)
                meta_data["value"].append(str(value))
        
        # Add API key status if available
        if metadata and "api_key_status" in metadata:
            meta_data["property"].append("API Key Status")
            meta_data["value"].append(metadata["api_key_status"])
        
        # Add subscription requirement if available
        if metadata and "requires_subscription" in metadata:
            meta_data["property"].append("Requires Subscription")
            meta_data["value"].append(str(metadata["requires_subscription"]))
        
        # Add data quality score if available
        if metadata and "data_quality_score" in metadata:
            meta_data["property"].append("Data Quality Score")
            meta_data["value"].append(str(metadata["data_quality_score"]))
        
        return pd.DataFrame(meta_data)


def export_to_excel(
    df: pd.DataFrame,
    filename: Optional[str] = None,
    site_id: Optional[str] = None,
    metadata: Optional[Dict[str, Any]] = None,
) -> Path:
    """
    Convenience function to export a DataFrame to Excel.
    
    Args:
        df: DataFrame to export
        filename: Output filename
        site_id: Site identifier
        metadata: Additional metadata
    
    Returns:
        Path to the created Excel file
    """
    exporter = ExcelExporter()
    return exporter.export(df, filename, site_id, metadata)

