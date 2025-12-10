"""
Analyze DG ECFIN Excel file structure
"""

import pandas as pd
import openpyxl

file_path = "FRED/exploration/dg_ecfin_samples/extracted/main_indicators_nace2.xlsx"

print("Analyzing DG ECFIN Main Indicators Excel File")
print("=" * 80)

# Load workbook to see sheet names
wb = openpyxl.load_workbook(file_path)
print(f"\nSheet names: {wb.sheetnames}\n")

# Read each sheet
for sheet_name in wb.sheetnames[:5]:  # First 5 sheets
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print('='*80)

    try:
        # Try reading without header first
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=20)
        print(f"Shape: {df.shape}")
        print("\nFirst 20 rows:")
        print(df.to_string())
    except Exception as e:
        print(f"Error reading sheet: {e}")

print("\n" + "=" * 80)
print("Looking for specific indicators...")
print("=" * 80)

# Try to find sheets with ESI, EEI, Consumer Confidence
indicators_to_find = ['ESI', 'EEI', 'Consumer', 'Confidence', 'Economic Sentiment']

for sheet_name in wb.sheetnames:
    for indicator in indicators_to_find:
        if indicator.lower() in sheet_name.lower():
            print(f"\n✓ Found '{indicator}' in sheet: {sheet_name}")

            # Read this sheet
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=30)
                print(f"  Shape: {df.shape}")
                print(f"\n  Sample data:")
                print(df.head(30).to_string())
            except Exception as e:
                print(f"  Error: {e}")

            break  # Only show first match per sheet
